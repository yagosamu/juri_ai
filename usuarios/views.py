from django.contrib.auth import authenticate
from django.contrib import auth, messages
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.messages import constants
from django.contrib import messages
from django.contrib.auth.models import User
from .models import Cliente, Documentos, ConfiguracaoWhatsApp, ConsentimentoLGPD, Processo, Prazo, AndamentoProcesso, TRIBUNAL_CHOICES, Honorario, Pagamento
from django.db.models import Sum
from ia.models import AnaliseJurisprudencia
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.conf import settings
import datetime
import os



def home(request):
    return render(request, 'home.html')


def privacidade(request):
    return render(request, 'privacidade.html')


def termos(request):
    return render(request, 'termos.html')


def cadastro(request):
    if request.method == 'GET':
        return render(request, 'cadastro.html')
    elif request.method == 'POST':
        username = request.POST.get('username')
        senha = request.POST.get('senha')
        confirmar_senha = request.POST.get('confirmar_senha')
        aceito_termos = request.POST.get('aceito_termos')

        if not aceito_termos:
            messages.add_message(request, constants.ERROR, 'Você precisa aceitar os Termos de Uso e a Política de Privacidade para criar uma conta.')
            return redirect('cadastro')

        if not senha == confirmar_senha:
            messages.add_message(request, constants.ERROR, 'Senha e confirmar senha não são iguais.')
            return redirect('cadastro')

        if len(senha) < 6:
            messages.add_message(request, constants.ERROR, 'Sua senha deve ter pelo meno 6 caracteres.')
            return redirect('cadastro')

        users = User.objects.filter(username=username)

        if users.exists():
            messages.add_message(request, constants.ERROR, 'Já existe um usuário com esse username.')
            return redirect('cadastro')

        user = User.objects.create_user(
            username=username,
            password=senha
        )

        ip = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR'))
        if ip and ',' in ip:
            ip = ip.split(',')[0].strip()
        ConsentimentoLGPD.objects.create(user=user, versao='1.0', ip=ip)

        return redirect('login')


def login(request):
    if request.method == 'GET':
        return render(request, 'login.html')
    elif request.method == 'POST':
        username = request.POST.get('username')
        senha = request.POST.get('senha')
        
        user = authenticate(request, username=username, password=senha)
        if user is not None:
            auth.login(request, user)
            return redirect('dashboard')
        else:
            messages.add_message(request, constants.ERROR, 'Usuário ou senha inválidos.')
            return redirect('login')


def _get_proximos_eventos():
    """Busca os próximos 5 eventos do Google Calendar. Retorna [] se não configurado."""
    try:
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build

        token_path = os.path.join(settings.BASE_DIR, 'token.json')
        if not os.path.exists(token_path):
            return []

        creds = Credentials.from_authorized_user_file(token_path)
        service = build('calendar', 'v3', credentials=creds)

        agora = datetime.datetime.utcnow().isoformat() + 'Z'
        resultado = service.events().list(
            calendarId='primary',
            timeMin=agora,
            maxResults=5,
            singleEvents=True,
            orderBy='startTime',
        ).execute()
        return resultado.get('items', [])
    except Exception:
        return []


@login_required
def dashboard(request):
    clientes_qs = Cliente.objects.filter(user=request.user)
    documentos_qs = Documentos.objects.filter(cliente__user=request.user)

    total_clientes    = clientes_qs.count()
    clientes_ativos   = clientes_qs.filter(status=True).count()
    clientes_inativos = clientes_qs.filter(status=False).count()

    total_documentos        = documentos_qs.count()
    documentos_processados  = documentos_qs.exclude(content='').count()

    alertas = (AnaliseJurisprudencia.objects
               .filter(documento__cliente__user=request.user,
                       classificacao__in=['Alto', 'Crítico'])
               .select_related('documento', 'documento__cliente')
               .order_by('-data_criacao')[:5])

    proximos_eventos = _get_proximos_eventos()

    hoje    = datetime.date.today()
    semana  = hoje + datetime.timedelta(days=7)
    prazos_semana = (Prazo.objects
                     .filter(user=request.user, status='pendente',
                             data_prazo__gte=hoje, data_prazo__lte=semana)
                     .select_related('processo')
                     .order_by('data_prazo'))

    return render(request, 'dashboard.html', {
        'total_clientes': total_clientes,
        'clientes_ativos': clientes_ativos,
        'clientes_inativos': clientes_inativos,
        'total_documentos': total_documentos,
        'documentos_processados': documentos_processados,
        'alertas': alertas,
        'proximos_eventos': proximos_eventos,
        'prazos_semana': prazos_semana,
        'hoje': hoje,
    })


@login_required
def clientes(request):
    if request.method == 'GET':
        clientes = Cliente.objects.filter(user=request.user)
        return render(request, 'clientes.html', {'clientes': clientes})
    elif request.method == 'POST':
        nome = request.POST.get('nome')
        email = request.POST.get('email')
        tipo = request.POST.get('tipo')
        status = request.POST.get('status') == 'on'
        
        Cliente.objects.create(
            nome=nome,
            email=email,
            tipo=tipo,
            status=status,
            user=request.user
        )
        
        messages.add_message(request, constants.SUCCESS, 'Cliente cadastrado com sucesso!')
        return redirect('clientes')
    

@login_required
def cliente(request, id):
    cliente = get_object_or_404(Cliente, id=id, user=request.user)
    if request.method == 'GET':
        documentos = Documentos.objects.filter(cliente=cliente)
        honorarios = Honorario.objects.filter(cliente=cliente, user=request.user).order_by('vencimento')
        tab = request.GET.get('tab', 'documentos')
        return render(request, 'cliente.html', {
            'cliente':    cliente,
            'documentos': documentos,
            'honorarios': honorarios,
            'tab':        tab,
        })
    elif request.method == 'POST':
        tipo = request.POST.get('tipo')
        documento = request.FILES.get('documento')
        data = request.POST.get('data')

        documentos = Documentos(
            cliente=cliente,
            tipo=tipo,
            arquivo=documento,
            data_upload=data
        )
        documentos.save()

        return redirect(reverse('cliente', kwargs={'id': cliente.id}))


@login_required
def configuracao_whatsapp(request):
    config = ConfiguracaoWhatsApp.objects.filter(user=request.user).first()
    if request.method == 'GET':
        return render(request, 'configuracao_whatsapp.html', {'config': config})
    elif request.method == 'POST':
        base_url = request.POST.get('base_url')
        instancia = request.POST.get('instancia')
        api_key = request.POST.get('api_key')

        if config:
            config.base_url = base_url
            config.instancia = instancia
            if api_key:
                config.api_key = api_key
            config.save()
        else:
            ConfiguracaoWhatsApp.objects.create(
                user=request.user,
                base_url=base_url,
                instancia=instancia,
                api_key=api_key,
            )

        messages.add_message(request, constants.SUCCESS, 'Configuração salva com sucesso!')
        return redirect('configuracao_whatsapp')


@login_required
def lista_processos(request):
    qs = Processo.objects.filter(user=request.user).select_related('cliente')

    # Filtros via querystring
    status_filtro   = request.GET.get('status', '')
    tribunal_filtro = request.GET.get('tribunal', '')

    if status_filtro:
        qs = qs.filter(status=status_filtro)
    if tribunal_filtro:
        qs = qs.filter(tribunal=tribunal_filtro)

    return render(request, 'processos.html', {
        'processos':        qs,
        'status_filtro':    status_filtro,
        'tribunal_filtro':  tribunal_filtro,
        'tribunal_choices': TRIBUNAL_CHOICES,
    })


@login_required
def criar_processo(request):
    clientes_usuario = Cliente.objects.filter(user=request.user)

    if request.method == 'GET':
        return render(request, 'criar_processo.html', {
            'clientes':         clientes_usuario,
            'tribunal_choices': TRIBUNAL_CHOICES,
        })

    # POST — valida e cria
    numero_cnj        = request.POST.get('numero_cnj', '').strip()
    tribunal          = request.POST.get('tribunal', '').strip()
    tribunal_outro    = request.POST.get('tribunal_outro', '').strip()
    vara              = request.POST.get('vara', '').strip()
    comarca           = request.POST.get('comarca', '').strip()
    tipo_acao         = request.POST.get('tipo_acao', '').strip()
    polo_ativo        = request.POST.get('polo_ativo', '').strip()
    polo_passivo      = request.POST.get('polo_passivo', '').strip()
    status            = request.POST.get('status', 'ativo')
    data_distribuicao = request.POST.get('data_distribuicao') or None
    valor_causa       = request.POST.get('valor_causa') or None
    cliente_id        = request.POST.get('cliente_id')

    # Validações
    digitos = numero_cnj.replace('-', '').replace('.', '')
    if len(digitos) != 20:
        messages.add_message(request, constants.ERROR, 'Número CNJ inválido. Informe os 20 dígitos (ex: 0000000-00.0000.0.00.0000).')
        return render(request, 'criar_processo.html', {
            'clientes':         clientes_usuario,
            'tribunal_choices': TRIBUNAL_CHOICES,
            'form_data':        request.POST,
        })

    if tribunal == 'outro' and not tribunal_outro:
        messages.add_message(request, constants.ERROR, 'Informe o nome do tribunal no campo "Tribunal (outro)".')
        return render(request, 'criar_processo.html', {
            'clientes':         clientes_usuario,
            'tribunal_choices': TRIBUNAL_CHOICES,
            'form_data':        request.POST,
        })

    cliente_obj = get_object_or_404(Cliente, id=cliente_id, user=request.user)

    if Processo.objects.filter(numero_cnj=numero_cnj, user=request.user).exists():
        messages.add_message(request, constants.ERROR, 'Você já cadastrou um processo com este número CNJ.')
        return render(request, 'criar_processo.html', {
            'clientes':         clientes_usuario,
            'tribunal_choices': TRIBUNAL_CHOICES,
            'form_data':        request.POST,
        })

    processo = Processo.objects.create(
        numero_cnj        = numero_cnj,
        tribunal          = tribunal,
        tribunal_outro    = tribunal_outro,
        vara              = vara,
        comarca           = comarca,
        tipo_acao         = tipo_acao,
        polo_ativo        = polo_ativo,
        polo_passivo      = polo_passivo,
        status            = status,
        data_distribuicao = data_distribuicao,
        valor_causa       = valor_causa,
        cliente           = cliente_obj,
        user              = request.user,
    )

    messages.add_message(request, constants.SUCCESS, 'Processo cadastrado com sucesso!')
    return redirect(reverse('processo', kwargs={'id': processo.id}))


@login_required
def processo(request, id):
    processo_obj = get_object_or_404(Processo, id=id, user=request.user)

    # Documentos vinculados ao processo
    docs_vinculados = Documentos.objects.filter(processo=processo_obj).select_related('cliente')

    # Documentos do mesmo cliente ainda sem processo (para vinculação)
    docs_disponiveis = Documentos.objects.filter(
        cliente=processo_obj.cliente,
        processo__isnull=True
    )

    # Análises de risco via documentos vinculados
    from ia.models import AnaliseJurisprudencia
    analises = AnaliseJurisprudencia.objects.filter(
        documento__processo=processo_obj
    ).select_related('documento').order_by('-data_criacao')

    # Prazos do processo
    prazos = Prazo.objects.filter(processo=processo_obj)

    # Andamentos DataJud
    andamentos = AndamentoProcesso.objects.filter(processo=processo_obj)

    # Honorários do processo
    honorarios = Honorario.objects.filter(processo=processo_obj, user=request.user).order_by('vencimento')

    # Tab ativa (querystring ?tab=documentos|analises|prazos|andamentos|honorarios)
    tab = request.GET.get('tab', 'documentos')

    return render(request, 'processo.html', {
        'processo':         processo_obj,
        'docs_vinculados':  docs_vinculados,
        'docs_disponiveis': docs_disponiveis,
        'analises':         analises,
        'prazos':           prazos,
        'andamentos':       andamentos,
        'honorarios':       honorarios,
        'tab':              tab,
    })


@login_required
def vincular_documento(request, id):
    """Vincula um documento existente do cliente ao processo via POST."""
    processo_obj = get_object_or_404(Processo, id=id, user=request.user)

    if request.method == 'POST':
        doc_id = request.POST.get('documento_id')
        doc = get_object_or_404(Documentos, id=doc_id, cliente=processo_obj.cliente)
        doc.processo = processo_obj
        doc.save()
        messages.add_message(request, constants.SUCCESS, 'Documento vinculado ao processo.')

    return redirect(reverse('processo', kwargs={'id': id}) + '?tab=documentos')


@login_required
def atualizar_datajud(request, id):
    """Dispara consulta DataJud assíncrona para o processo."""
    processo_obj = get_object_or_404(Processo, id=id, user=request.user)

    if request.method == 'POST':
        if processo_obj.tribunal == 'outro':
            messages.add_message(request, constants.ERROR,
                                 'Processos com tribunal "Outro" não podem ser consultados no DataJud.')
        else:
            from django_q.tasks import async_task
            async_task('ia.tasks.consultar_datajud', processo_obj.id)
            messages.add_message(request, constants.SUCCESS,
                                 'Consulta ao DataJud iniciada. Os andamentos aparecerão em instantes.')

    return redirect(reverse('processo', kwargs={'id': id}) + '?tab=andamentos')


@login_required
def editar_processo(request, id):
    processo_obj    = get_object_or_404(Processo, id=id, user=request.user)
    clientes_usuario = Cliente.objects.filter(user=request.user)

    if request.method == 'GET':
        return render(request, 'editar_processo.html', {
            'processo':         processo_obj,
            'clientes':         clientes_usuario,
            'tribunal_choices': TRIBUNAL_CHOICES,
        })

    # POST — atualiza
    numero_cnj        = request.POST.get('numero_cnj', '').strip()
    tribunal          = request.POST.get('tribunal', '').strip()
    tribunal_outro    = request.POST.get('tribunal_outro', '').strip()
    vara              = request.POST.get('vara', '').strip()
    comarca           = request.POST.get('comarca', '').strip()
    tipo_acao         = request.POST.get('tipo_acao', '').strip()
    polo_ativo        = request.POST.get('polo_ativo', '').strip()
    polo_passivo      = request.POST.get('polo_passivo', '').strip()
    status            = request.POST.get('status', processo_obj.status)
    data_distribuicao = request.POST.get('data_distribuicao') or None
    valor_causa       = request.POST.get('valor_causa') or None
    cliente_id        = request.POST.get('cliente_id')

    digitos = numero_cnj.replace('-', '').replace('.', '')
    if len(digitos) != 20:
        messages.add_message(request, constants.ERROR, 'Número CNJ inválido.')
        return render(request, 'editar_processo.html', {
            'processo':         processo_obj,
            'clientes':         clientes_usuario,
            'tribunal_choices': TRIBUNAL_CHOICES,
        })

    if tribunal == 'outro' and not tribunal_outro:
        messages.add_message(request, constants.ERROR, 'Informe o nome do tribunal no campo "Tribunal (outro)".')
        return render(request, 'editar_processo.html', {
            'processo':         processo_obj,
            'clientes':         clientes_usuario,
            'tribunal_choices': TRIBUNAL_CHOICES,
        })

    # Verifica unicidade excluindo o próprio processo
    if (Processo.objects.filter(numero_cnj=numero_cnj, user=request.user)
                        .exclude(id=processo_obj.id).exists()):
        messages.add_message(request, constants.ERROR, 'Você já tem outro processo com este número CNJ.')
        return render(request, 'editar_processo.html', {
            'processo':         processo_obj,
            'clientes':         clientes_usuario,
            'tribunal_choices': TRIBUNAL_CHOICES,
        })

    cliente_obj = get_object_or_404(Cliente, id=cliente_id, user=request.user)

    processo_obj.numero_cnj        = numero_cnj
    processo_obj.tribunal          = tribunal
    processo_obj.tribunal_outro    = tribunal_outro
    processo_obj.vara              = vara
    processo_obj.comarca           = comarca
    processo_obj.tipo_acao         = tipo_acao
    processo_obj.polo_ativo        = polo_ativo
    processo_obj.polo_passivo      = polo_passivo
    processo_obj.status            = status
    processo_obj.data_distribuicao = data_distribuicao
    processo_obj.valor_causa       = valor_causa
    processo_obj.cliente           = cliente_obj
    processo_obj.save()

    messages.add_message(request, constants.SUCCESS, 'Processo atualizado com sucesso!')
    return redirect(reverse('processo', kwargs={'id': processo_obj.id}))


@login_required
def arquivar_processo(request, id):
    """Soft delete: muda status para 'arquivado'."""
    if request.method != 'POST':
        return redirect('lista_processos')

    processo_obj = get_object_or_404(Processo, id=id, user=request.user)
    processo_obj.status = 'arquivado'
    processo_obj.save()

    messages.add_message(request, constants.SUCCESS,
                         f'Processo {processo_obj} arquivado.')
    return redirect('lista_processos')


@login_required
def lista_prazos(request):
    """Agenda — prazos dos próximos 30 dias."""
    hoje  = datetime.date.today()
    limite = hoje + datetime.timedelta(days=30)

    tipo_filtro   = request.GET.get('tipo', '')
    status_filtro = request.GET.get('status', 'pendente')

    qs = Prazo.objects.filter(
        user=request.user,
        data_prazo__gte=hoje,
        data_prazo__lte=limite,
    ).select_related('processo', 'processo__cliente')

    if tipo_filtro:
        qs = qs.filter(tipo=tipo_filtro)
    if status_filtro:
        qs = qs.filter(status=status_filtro)

    return render(request, 'prazos.html', {
        'prazos':        qs,
        'tipo_filtro':   tipo_filtro,
        'status_filtro': status_filtro,
        'tipo_choices':  Prazo.TIPO_CHOICES,
        'hoje':          hoje,
        'limite':        limite,
    })


@login_required
def criar_prazo(request, processo_id):
    processo_obj = get_object_or_404(Processo, id=processo_id, user=request.user)

    if request.method == 'GET':
        return render(request, 'criar_prazo.html', {
            'processo':    processo_obj,
            'tipo_choices': Prazo.TIPO_CHOICES,
        })

    descricao         = request.POST.get('descricao', '').strip()
    data_prazo        = request.POST.get('data_prazo', '').strip()
    tipo              = request.POST.get('tipo', '').strip()
    alerta_dias_antes = request.POST.get('alerta_dias_antes', '3').strip()

    if not descricao or not data_prazo or not tipo:
        messages.add_message(request, constants.ERROR, 'Preencha todos os campos obrigatórios.')
        return render(request, 'criar_prazo.html', {
            'processo':    processo_obj,
            'tipo_choices': Prazo.TIPO_CHOICES,
            'form_data':   request.POST,
        })

    prazo = Prazo.objects.create(
        descricao         = descricao,
        data_prazo        = data_prazo,
        tipo              = tipo,
        processo          = processo_obj,
        alerta_dias_antes = int(alerta_dias_antes) if alerta_dias_antes.isdigit() else 3,
        user              = request.user,
    )

    from usuarios.calendar_utils import criar_evento_prazo
    event_id = criar_evento_prazo(prazo)
    if event_id:
        prazo.google_event_id = event_id
        prazo.save(update_fields=['google_event_id'])

    messages.add_message(request, constants.SUCCESS, 'Prazo cadastrado com sucesso!')
    return redirect(reverse('processo', kwargs={'id': processo_id}) + '?tab=prazos')


@login_required
def editar_prazo(request, id):
    prazo = get_object_or_404(Prazo, id=id, user=request.user)

    if request.method == 'GET':
        return render(request, 'editar_prazo.html', {
            'prazo':        prazo,
            'tipo_choices': Prazo.TIPO_CHOICES,
        })

    descricao         = request.POST.get('descricao', '').strip()
    data_prazo        = request.POST.get('data_prazo', '').strip()
    tipo              = request.POST.get('tipo', '').strip()
    alerta_dias_antes = request.POST.get('alerta_dias_antes', str(prazo.alerta_dias_antes)).strip()

    if not descricao or not data_prazo or not tipo:
        messages.add_message(request, constants.ERROR, 'Preencha todos os campos obrigatórios.')
        return render(request, 'editar_prazo.html', {
            'prazo':        prazo,
            'tipo_choices': Prazo.TIPO_CHOICES,
        })

    prazo.descricao         = descricao
    prazo.data_prazo        = data_prazo
    prazo.tipo              = tipo
    prazo.alerta_dias_antes = int(alerta_dias_antes) if alerta_dias_antes.isdigit() else prazo.alerta_dias_antes
    prazo.save()

    from usuarios.calendar_utils import atualizar_evento_prazo
    atualizar_evento_prazo(prazo)

    messages.add_message(request, constants.SUCCESS, 'Prazo atualizado com sucesso!')
    return redirect(reverse('processo', kwargs={'id': prazo.processo.id}) + '?tab=prazos')


@login_required
def concluir_prazo(request, id):
    if request.method != 'POST':
        return redirect('lista_prazos')

    prazo = get_object_or_404(Prazo, id=id, user=request.user)
    prazo.status = 'concluido'
    prazo.save(update_fields=['status'])

    from usuarios.calendar_utils import atualizar_evento_prazo
    atualizar_evento_prazo(prazo)

    messages.add_message(request, constants.SUCCESS, f'Prazo "{prazo.descricao}" marcado como concluído.')
    next_url = request.POST.get('next', '')
    if next_url and next_url.startswith('/'):
        return redirect(next_url)
    return redirect(reverse('processo', kwargs={'id': prazo.processo.id}) + '?tab=prazos')


@login_required
def cancelar_prazo(request, id):
    if request.method != 'POST':
        return redirect('lista_prazos')

    prazo = get_object_or_404(Prazo, id=id, user=request.user)
    prazo.status = 'cancelado'
    prazo.save(update_fields=['status'])

    from usuarios.calendar_utils import cancelar_evento_prazo
    cancelar_evento_prazo(prazo)

    messages.add_message(request, constants.SUCCESS, f'Prazo "{prazo.descricao}" cancelado.')
    next_url = request.POST.get('next', '')
    if next_url and next_url.startswith('/'):
        return redirect(next_url)
    return redirect(reverse('processo', kwargs={'id': prazo.processo.id}) + '?tab=prazos')


@login_required
def excluir_conta(request):
    if request.method == 'GET':
        return render(request, 'excluir_conta.html')

    senha_confirmacao = request.POST.get('senha_confirmacao')
    user = authenticate(request, username=request.user.username, password=senha_confirmacao)

    if user is None:
        messages.add_message(request, constants.ERROR, 'Senha incorreta. A conta não foi excluída.')
        return redirect('excluir_conta')

    # Tenta limpar registros no LanceDB silenciosamente antes de deletar o User
    try:
        import lancedb
        from django.conf import settings as django_settings
        import os

        lancedb_path = os.path.join(django_settings.BASE_DIR, 'lancedb')
        db = lancedb.connect(lancedb_path)

        cliente_ids = list(
            Cliente.objects.filter(user=user).values_list('id', flat=True)
        )
        if cliente_ids and 'documentos' in db.table_names():
            table = db.open_table('documentos')
            ids_str = ', '.join(str(cid) for cid in cliente_ids)
            table.delete(f"cliente_id IN ({ids_str})")
    except Exception:
        pass  # Falha no LanceDB não bloqueia a exclusão da conta

    # Deleta dados do usuário (CASCADE cobre: Cliente → Documentos, ConfiguracaoWhatsApp, ConsentimentoLGPD)
    auth.logout(request)
    user.delete()

    messages.add_message(request, constants.SUCCESS, 'Sua conta foi excluída com sucesso.')
    return redirect('home')


# ── Financeiro ───────────────────────────────────────────────────────────────

@login_required
def lista_financeiro(request):
    hoje       = datetime.date.today()
    inicio_mes = hoje.replace(day=1)

    honorarios_qs = (Honorario.objects
                     .filter(user=request.user)
                     .select_related('cliente', 'processo'))

    status_filtro = request.GET.get('status', '')
    tipo_filtro   = request.GET.get('tipo', '')
    if status_filtro:
        honorarios_qs = honorarios_qs.filter(status=status_filtro)
    if tipo_filtro:
        honorarios_qs = honorarios_qs.filter(tipo=tipo_filtro)

    base_qs = Honorario.objects.filter(user=request.user)

    receita_mes = (Pagamento.objects
                   .filter(honorario__user=request.user,
                           data_pagamento__gte=inicio_mes,
                           data_pagamento__lte=hoje)
                   .aggregate(total=Sum('valor_pago'))['total'] or 0)

    a_receber = (base_qs
                 .filter(status='pendente', vencimento__gte=hoje)
                 .aggregate(total=Sum('valor_total'))['total'] or 0)

    em_atraso = (base_qs
                 .filter(status='pendente', vencimento__lt=hoje)
                 .aggregate(total=Sum('valor_total'))['total'] or 0)

    return render(request, 'financeiro.html', {
        'honorarios':     honorarios_qs,
        'receita_mes':    receita_mes,
        'a_receber':      a_receber,
        'em_atraso':      em_atraso,
        'status_filtro':  status_filtro,
        'tipo_filtro':    tipo_filtro,
        'status_choices': Honorario.STATUS_CHOICES,
        'tipo_choices':   Honorario.TIPO_CHOICES,
        'hoje':           hoje,
    })


@login_required
def criar_honorario(request):
    clientes_qs  = Cliente.objects.filter(user=request.user)
    processos_qs = Processo.objects.filter(user=request.user).select_related('cliente')

    # Pré-vinculação via querystring (?cliente_id=X&processo_id=Y)
    cliente_id_pre  = request.GET.get('cliente_id') or request.POST.get('cliente_id_pre')
    processo_id_pre = request.GET.get('processo_id') or request.POST.get('processo_id_pre')

    if request.method == 'GET':
        return render(request, 'criar_honorario.html', {
            'clientes':        clientes_qs,
            'processos':       processos_qs,
            'cliente_id_pre':  cliente_id_pre,
            'processo_id_pre': processo_id_pre,
            'tipo_choices':    Honorario.TIPO_CHOICES,
        })

    descricao   = request.POST.get('descricao', '').strip()
    valor_total = request.POST.get('valor_total', '').strip()
    tipo        = request.POST.get('tipo', '').strip()
    vencimento  = request.POST.get('vencimento', '').strip()
    cliente_id  = request.POST.get('cliente_id')
    processo_id = request.POST.get('processo_id') or None

    if not all([descricao, valor_total, tipo, vencimento, cliente_id]):
        messages.add_message(request, constants.ERROR, 'Preencha todos os campos obrigatórios.')
        return render(request, 'criar_honorario.html', {
            'clientes':     clientes_qs,
            'processos':    processos_qs,
            'tipo_choices': Honorario.TIPO_CHOICES,
            'form_data':    request.POST,
        })

    cliente_obj  = get_object_or_404(Cliente, id=cliente_id, user=request.user)
    processo_obj = None
    if processo_id:
        processo_obj = get_object_or_404(Processo, id=processo_id, user=request.user)

    Honorario.objects.create(
        cliente     = cliente_obj,
        processo    = processo_obj,
        descricao   = descricao,
        valor_total = valor_total,
        tipo        = tipo,
        vencimento  = vencimento,
        user        = request.user,
    )
    messages.add_message(request, constants.SUCCESS, 'Honorário cadastrado com sucesso!')
    return redirect('lista_financeiro')


@login_required
def editar_honorario(request, id):
    honorario    = get_object_or_404(Honorario, id=id, user=request.user)
    clientes_qs  = Cliente.objects.filter(user=request.user)
    processos_qs = Processo.objects.filter(user=request.user).select_related('cliente')

    if request.method == 'GET':
        return render(request, 'editar_honorario.html', {
            'honorario':    honorario,
            'clientes':     clientes_qs,
            'processos':    processos_qs,
            'tipo_choices': Honorario.TIPO_CHOICES,
        })

    descricao   = request.POST.get('descricao', '').strip()
    valor_total = request.POST.get('valor_total', '').strip()
    tipo        = request.POST.get('tipo', '').strip()
    vencimento  = request.POST.get('vencimento', '').strip()
    cliente_id  = request.POST.get('cliente_id')
    processo_id = request.POST.get('processo_id') or None
    status      = request.POST.get('status', honorario.status)

    if not all([descricao, valor_total, tipo, vencimento, cliente_id]):
        messages.add_message(request, constants.ERROR, 'Preencha todos os campos obrigatórios.')
        return render(request, 'editar_honorario.html', {
            'honorario':    honorario,
            'clientes':     clientes_qs,
            'processos':    processos_qs,
            'tipo_choices': Honorario.TIPO_CHOICES,
        })

    honorario.cliente     = get_object_or_404(Cliente, id=cliente_id, user=request.user)
    honorario.processo    = get_object_or_404(Processo, id=processo_id, user=request.user) if processo_id else None
    honorario.descricao   = descricao
    honorario.valor_total = valor_total
    honorario.tipo        = tipo
    honorario.vencimento  = vencimento
    honorario.status      = status
    honorario.save()

    messages.add_message(request, constants.SUCCESS, 'Honorário atualizado com sucesso!')
    return redirect('lista_financeiro')


@login_required
def marcar_pago(request, id):
    if request.method != 'POST':
        return redirect('lista_financeiro')

    honorario      = get_object_or_404(Honorario, id=id, user=request.user)
    valor_pago     = request.POST.get('valor_pago', '').strip()
    data_pagamento = request.POST.get('data_pagamento', '').strip()
    observacao     = request.POST.get('observacao', '').strip()

    if not valor_pago or not data_pagamento:
        messages.add_message(request, constants.ERROR, 'Informe o valor pago e a data.')
        return redirect('lista_financeiro')

    Pagamento.objects.create(
        honorario      = honorario,
        valor_pago     = valor_pago,
        data_pagamento = data_pagamento,
        observacao     = observacao,
    )
    # Honorario.status atualizado automaticamente pelo Pagamento.save()
    messages.add_message(request, constants.SUCCESS, 'Pagamento registrado com sucesso!')

    next_url = request.POST.get('next', '')
    if next_url and next_url.startswith('/'):
        return redirect(next_url)
    return redirect('lista_financeiro')


@login_required
def cancelar_honorario(request, id):
    if request.method != 'POST':
        return redirect('lista_financeiro')

    honorario = get_object_or_404(Honorario, id=id, user=request.user)
    honorario.status = 'cancelado'
    honorario.save(update_fields=['status'])

    messages.add_message(request, constants.SUCCESS,
                         f'Honorário "{honorario.descricao}" cancelado.')
    next_url = request.POST.get('next', '')
    if next_url and next_url.startswith('/'):
        return redirect(next_url)
    return redirect('lista_financeiro')


@login_required
def relatorio_financeiro(request):
    hoje        = datetime.date.today()
    data_inicio = (request.POST.get('data_inicio') or
                   request.GET.get('data_inicio') or
                   hoje.replace(day=1).isoformat())
    data_fim    = (request.POST.get('data_fim') or
                   request.GET.get('data_fim') or
                   hoje.isoformat())

    try:
        data_inicio_obj = datetime.date.fromisoformat(data_inicio)
        data_fim_obj    = datetime.date.fromisoformat(data_fim)
    except ValueError:
        data_inicio_obj = hoje.replace(day=1)
        data_fim_obj    = hoje

    honorarios = (Honorario.objects
                  .filter(user=request.user,
                          vencimento__gte=data_inicio_obj,
                          vencimento__lte=data_fim_obj)
                  .select_related('cliente', 'processo')
                  .order_by('vencimento'))

    total_geral    = honorarios.aggregate(total=Sum('valor_total'))['total'] or 0
    total_recebido = (Pagamento.objects
                     .filter(honorario__user=request.user,
                             data_pagamento__gte=data_inicio_obj,
                             data_pagamento__lte=data_fim_obj)
                     .aggregate(total=Sum('valor_pago'))['total'] or 0)
    total_pendente = (honorarios.filter(status='pendente')
                     .aggregate(total=Sum('valor_total'))['total'] or 0)

    formato = request.POST.get('formato', '')

    if request.method == 'POST' and formato == 'pdf':
        from io import BytesIO
        from django.http import HttpResponse
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm

        buffer = BytesIO()
        doc    = SimpleDocTemplate(buffer, pagesize=A4,
                                   rightMargin=2*cm, leftMargin=2*cm,
                                   topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        story  = []

        story.append(Paragraph('Relatório de Fluxo de Caixa — JuriAI', styles['Title']))
        story.append(Paragraph(
            f'Período: {data_inicio_obj.strftime("%d/%m/%Y")} a {data_fim_obj.strftime("%d/%m/%Y")}',
            styles['Normal'],
        ))
        story.append(Spacer(1, 0.5*cm))

        dados = [['Vencimento', 'Cliente', 'Descrição', 'Tipo', 'Valor (R$)', 'Status']]
        for h in honorarios:
            dados.append([
                h.vencimento.strftime('%d/%m/%Y'),
                h.cliente.nome,
                h.descricao,
                h.get_tipo_display(),
                f'{h.valor_total:,.2f}',
                h.get_status_display(),
            ])
        dados.append(['', '', '', 'TOTAL', f'{total_geral:,.2f}', ''])

        tabela = Table(dados, repeatRows=1)
        tabela.setStyle(TableStyle([
            ('BACKGROUND',     (0, 0),  (-1, 0),  colors.HexColor('#1e293b')),
            ('TEXTCOLOR',      (0, 0),  (-1, 0),  colors.white),
            ('FONTNAME',       (0, 0),  (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',       (0, 0),  (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1),  (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
            ('BACKGROUND',     (0, -1), (-1, -1), colors.HexColor('#f1f5f9')),
            ('FONTNAME',       (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID',           (0, 0),  (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('ALIGN',          (4, 0),  (4, -1),  'RIGHT'),
            ('VALIGN',         (0, 0),  (-1, -1), 'MIDDLE'),
            ('PADDING',        (0, 0),  (-1, -1), 6),
        ]))
        story.append(tabela)
        doc.build(story)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="fluxo_caixa_{data_inicio}_{data_fim}.pdf"'
        )
        return response

    if request.method == 'POST' and formato == 'excel':
        from io import BytesIO
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Fluxo de Caixa'

        ws.append(['Vencimento', 'Cliente', 'Descrição', 'Tipo', 'Valor (R$)', 'Status'])
        fill       = PatternFill(fill_type='solid', fgColor='1e293b')
        bold_white = Font(color='FFFFFF', bold=True)
        for cell in ws[1]:
            cell.fill      = fill
            cell.font      = bold_white
            cell.alignment = Alignment(horizontal='center')

        for h in honorarios:
            ws.append([
                h.vencimento.strftime('%d/%m/%Y'),
                h.cliente.nome,
                h.descricao,
                h.get_tipo_display(),
                float(h.valor_total),
                h.get_status_display(),
            ])

        ws.append(['', '', '', 'TOTAL', float(total_geral), ''])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        for col in ws.columns:
            width = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = width + 4

        buffer = BytesIO()
        wb.save(buffer)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="fluxo_caixa_{data_inicio}_{data_fim}.xlsx"'
        )
        return response

    return render(request, 'relatorio_financeiro.html', {
        'honorarios':     honorarios,
        'total_geral':    total_geral,
        'total_recebido': total_recebido,
        'total_pendente': total_pendente,
        'data_inicio':    data_inicio,
        'data_fim':       data_fim,
    })


@login_required
def logout(request):
    auth.logout(request)
    return redirect('login')
